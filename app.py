
import io
import time
import requests
from openpyxl import load_workbook
import streamlit as st
from streamlit_autorefresh import st_autorefresh

st.set_page_config(page_title="HSE XLS Monitor", layout="centered")

DEFAULT_URL = "https://priem44.hse.ru/ABITREPORTS/MAGREPORTS/EnrollmentList/28367398628_Commercial.xlsx"

st.title("Монитор ВШЭ: АБД")
st.caption("Ниже отображается суммарное количество заключенных и оплаченных договоров. Если ввести регистрационному номер абитуриента, то ниже оборажается его суммарный бал а так же место в соответствующих рейтингах. Кешированные результаты обновляются раз в час.")

# url = st.text_input("XLS(X) URL", value=DEFAULT_URL)
url = DEFAULT_URL

# Auto-refresh every hour
st_autorefresh(interval=60 * 60 * 1000, key="hourly_refresh")

def is_yes(val):
    if val is None:
        return False
    return str(val).strip().lower() == "да"

def norm_reg(v):
    """Return a tuple: (raw_str, canonical_digits) for comparison that ignores leading zeros and spaces."""
    if v is None:
        return None, None
    s = str(v).strip()
    digits = "".join(ch for ch in s if ch.isdigit())
    digits_nz = digits.lstrip("0") or "0" if digits else None
    return s, digits_nz

def to_number(val):
    """Parse numbers that may come as strings with comma decimal. Return float or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "").replace("\u00A0", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

@st.cache_data(ttl=60*60, show_spinner=True)
def fetch_and_parse(u: str):
    r = requests.get(u, timeout=60)
    r.raise_for_status()
    bio = io.BytesIO(r.content)
    wb = load_workbook(bio, data_only=True)
    ws = wb.active  # first sheet

    a20 = ws["A20"].value

    # Detect columns on header row (21)
    header_row = 21
    reg_col_idx = 2  # fallback B
    score_col_idx = 6  # fallback F
    try:
        for ci in range(1, 40):
            val = ws.cell(row=header_row, column=ci).value
            if not val:
                continue
            sval = str(val).strip().lower()
            if sval == "регистрационный номер":
                reg_col_idx = ci
            if sval == "сумма конкурсных баллов":
                score_col_idx = ci
    except Exception:
        pass

    rows = []
    contracts = 0
    paid = 0
    for row in range(22, 501):
        reg_val = ws.cell(row=row, column=reg_col_idx).value
        raw, canon = norm_reg(reg_val)
        c_yes = is_yes(ws[f"H{row}"].value)
        p_yes = is_yes(ws[f"I{row}"].value)
        score_val = to_number(ws.cell(row=row, column=score_col_idx).value)
        if c_yes:
            contracts += 1
        if p_yes:
            paid += 1
        rows.append({
            "row": row,
            "reg_raw": raw,
            "reg_canon": canon,
            "contract": c_yes,
            "paid": p_yes,
            "score": score_val
        })

    contract_rows = [r for r in rows if r["contract"]]
    paid_rows = [r for r in rows if r["paid"]]

    return {
        "a20": a20,
        "contracts": contracts,
        "paid": paid,
        "ts": int(time.time()),
        "rows": rows,
        "contract_rows": contract_rows,
        "paid_rows": paid_rows
    }

def fmt_score(x):
    if x is None:
        return "—"
    if abs(x - round(x)) < 1e-9:
        return f"{int(round(x))}"
    return f"{x:.2f}"

try:
    data = fetch_and_parse(url)

    st.caption(f"**A20**: {data["a20"] or "—"}")

    c1, c2, c3 = st.columns(3)
    c1.metric("Заключенных договоров", data["contracts"])
    c2.metric("Оплаченных договоров", data["paid"])
    # c3.metric("Last check (UTC)", time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(data["ts"])))

    reg_input = st.text_input("", value="", placeholder="Введи сюда свой регистрационный номер")
    
    if reg_input.strip():
        reg_raw, reg_can = norm_reg(reg_input)
        found = None
        for r in data["rows"]:
            if r["reg_raw"] == reg_raw or (reg_can and r["reg_canon"] == reg_can):
                found = r
                break

        if not found:
            st.error("Регистрационный номер не найден.")
        else:
            st.metric("Сумма конкурсных баллов", fmt_score(found.get("score")))

            contract_rank = None
            if found["contract"]:
                for i, rr in enumerate(data["contract_rows"]):
                    if rr["row"] == found["row"]:
                        contract_rank = i + 1
                        break

            paid_rank = None
            if found["paid"]:
                for i, rr in enumerate(data["paid_rows"]):
                    if rr["row"] == found["row"]:
                        paid_rank = i + 1
                        break

            c1, c2 = st.columns(2)
            c1.metric("Место среди договоров", contract_rank if contract_rank is not None else "—")
            c2.metric("Место среди оплат", paid_rank if paid_rank is not None else "—")

            notes = []
            if contract_rank is None:
                notes.append("абитуриент не в списке **заключивших договор**")
            if paid_rank is None:
                notes.append("абитуриент не в списке **оплативших**")
            if notes:
                st.caption("; ".join(notes))

except Exception as e:
    st.error(f"Error: {e}")
    st.stop()

# with st.expander("Raw debug"):
#     st.code("App loaded. Enter reg number to see score and ranks.", language="text")
